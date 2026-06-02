import os
import io
import logging
import email
from email import policy
import gspread
from imapclient import IMAPClient
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
import json
import re

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# === ПЕРЕМЕННЫЕ ОКРУЖЕНИЯ ===
YANDEX_LOGIN = os.getenv("YANDEX_LOGIN")
YANDEX_APP_PASSWORD = os.getenv("YANDEX_APP_PASSWORD")
SENDER_TO_CHECK = os.getenv("SENDER_TO_CHECK")
SERVICE_ACCOUNT_JSON = os.getenv("SERVICE_ACCOUNT_JSON")
SPREADSHEET_ID = os.getenv("SPREADSHEET_ID")
WORKSHEET_NAME = os.getenv("WORKSHEET_NAME")

if not all([YANDEX_LOGIN, YANDEX_APP_PASSWORD, SENDER_TO_CHECK, SERVICE_ACCOUNT_JSON, SPREADSHEET_ID, WORKSHEET_NAME]):
    logging.error("Не все переменные окружения установлены")
    exit(1)

service_account_info = json.loads(SERVICE_ACCOUNT_JSON)

def extract_rfq_from_filename(filename: str) -> str:
    base = filename.split('_')[0].split('.')[0]
    match = re.search(r'([A-Z0-9]+)', base, re.IGNORECASE)
    return match.group(1) if match else base

def parse_quotation_excel(file_content: bytes, filename: str) -> list:
    try:
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        if "Quotation" not in wb.sheetnames:
            logging.error(f"Нет листа 'Quotation': {wb.sheetnames}")
            return []
        ws = wb["Quotation"]

        rfq_cell = ws["H2"].value
        rfq = str(rfq_cell).strip() if rfq_cell else extract_rfq_from_filename(filename)
        date_cell = ws["O2"].value
        if hasattr(date_cell, 'strftime'):
            date_str = date_cell.strftime("%d.%m.%Y")
        else:
            date_str = str(date_cell) if date_cell else ""

        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == "№":
                header_row = r
                break
        if not header_row:
            logging.error("Строка с '№' не найдена")
            return []

        rows = []
        row = header_row + 1
        while row <= ws.max_row:
            first_val = ws.cell(row, 1).value
            if not first_val or str(first_val).strip() == "":
                break
            qty = ws.cell(row, 6).value
            if isinstance(qty, (int, float)):
                qty_str = str(int(qty)) if qty == int(qty) else str(qty)
            else:
                qty_str = str(qty) if qty else ""
            target = ws.cell(row, 9).value
            if hasattr(target, 'strftime'):
                target_str = target.strftime("%d.%m.%Y")
            else:
                target_str = str(target) if target else ""

            rows.append({
                "RFQ": rfq,
                "Date": date_str,
                "№": str(first_val).split('.')[0],
                "PN": ws.cell(row, 2).value or "",
                "DESC": ws.cell(row, 4).value or "",
                "Alt": ws.cell(row, 5).value or "no",
                "R. Qty.": qty_str,
                "Unit": ws.cell(row, 7).value or "",
                "Req.Condition": ws.cell(row, 8).value or "",
                "Target Date": target_str,
                "Comment": ws.cell(row, 20).value or ""
            })
            row += 1
        return rows
    except Exception as e:
        logging.error(f"Ошибка парсинга Excel: {e}", exc_info=True)
        return []

def main():
    logging.info("Запуск")
    try:
        with IMAPClient("imap.yandex.ru") as client:
            client.login(YANDEX_LOGIN, YANDEX_APP_PASSWORD)
            client.select_folder("INBOX")
            messages = client.search(['UNSEEN', 'FROM', SENDER_TO_CHECK])
            logging.info(f"Новых писем: {len(messages)}")
            if not messages:
                return

            creds = Credentials.from_service_account_info(
                service_account_info,
                scopes=['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
            )
            gc = gspread.authorize(creds)
            sh = gc.open_by_key(SPREADSHEET_ID)
            ws = sh.worksheet(WORKSHEET_NAME)

            # Собираем существующие ключи (RFQ, №, PN) для проверки дубликатов
            all_rows = ws.get_all_values()
            existing_keys = set()
            if len(all_rows) > 1:
                for row in all_rows[1:]:
                    if len(row) >= 4:
                        rfq = row[0] if row[0] else ""
                        num = row[2] if len(row) > 2 else ""
                        pn = row[3] if len(row) > 3 else ""
                        if rfq and num and pn:
                            existing_keys.add((rfq, num, pn))

            # Заголовки, если лист пустой
            expected_headers = ["RFQ","Date","№","PN","DESC","Alt","R. Qty.","Unit","Req.Condition","Target Date","Comment"]
            if not all_rows:
                ws.append_row(expected_headers)
                logging.info("Заголовки добавлены")

            # Определяем соответствие колонок по первой строке
            headers_row = ws.row_values(1)
            col_index = {}
            for idx, val in enumerate(headers_row, start=1):
                if val in expected_headers:
                    col_index[val] = idx
            if "RFQ" not in col_index:
                logging.error("Заголовок 'RFQ' не найден в первой строке")
                return
            max_col = max(col_index.values()) if col_index else len(expected_headers)

            # Обрабатываем каждое письмо
            for msg_id in messages:
                logging.info(f"Обработка письма {msg_id}")
                email_data = client.fetch([msg_id], ['RFC822'])
                raw = email_data[msg_id][b'RFC822']
                msg = email.message_from_bytes(raw, policy=policy.default)

                xlsx, fname = None, None
                for part in msg.iter_attachments():
                    fname = part.get_filename()
                    if fname and fname.lower().endswith('.xlsx'):
                        xlsx = part.get_content()
                        break
                if not xlsx:
                    logging.warning("Нет вложения .xlsx")
                    continue

                data = parse_quotation_excel(xlsx, fname)
                if not data:
                    continue

                new_count = 0
                for row_data in data:
                    key = (row_data["RFQ"], row_data["№"], row_data["PN"])
                    if key in existing_keys:
                        logging.info(f"Пропущен дубликат: {key}")
                        continue
                    new_row = [""] * max_col
                    for header, value in row_data.items():
                        if header in col_index:
                            new_row[col_index[header] - 1] = value
                    ws.append_row(new_row)
                    existing_keys.add(key)
                    new_count += 1
                logging.info(f"Добавлено {new_count} новых строк")

    except Exception as e:
        logging.error(f"Ошибка: {e}", exc_info=True)

if __name__ == "__main__":
    main()
